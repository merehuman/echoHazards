"""
NRC (National Response Center) ingestor.

The NRC is the sole federal point of contact for reporting oil and chemical spills.
All incidents must be reported by law, making this the most comprehensive US spill dataset.

Data source:
  https://nrc.uscg.mil/FOIAdata.aspx
  Yearly Excel files from 1990-present. Place in NRC_DATA_DIR named nrc_YYYY.xlsx.

File structure (multi-sheet relational, all joined on SEQNOS):
  CALLS             — base record: date, responsible company, state
  INCIDENT_COMMONS  — location, coordinates (DMS format), description
  MATERIAL_INVOLVED — what was released and how much
  (other sheets contain railroad/vessel/vehicle detail we don't use)

Coordinates in INCIDENT_COMMONS are stored as degrees/minutes/seconds across
separate columns (LAT_DEG, LAT_MIN, LAT_SEC, LAT_QUAD) — we convert to decimal.
"""

import logging
import uuid
from datetime import datetime, date
from pathlib import Path
from typing import Iterator

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from models.incident import IncidentSource, IncidentType, Medium
from ingestors.base import BaseIngestor

logger = logging.getLogger(__name__)


def _parse_float(val) -> float | None:
    if val is None or val == "":
        return None
    try:
        return float(str(val).replace(",", "").strip())
    except (ValueError, TypeError):
        return None


def _parse_date(val) -> date | None:
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    raw = str(val).strip()
    for fmt in ("%m/%d/%Y %H:%M", "%m/%d/%Y", "%Y-%m-%d", "%m-%d-%Y", "%d-%b-%Y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def _dms_to_decimal(deg, minutes, sec, quad) -> float | None:
    """
    Convert degrees/minutes/seconds + quadrant to decimal degrees.
    LAT_QUAD: N/S, LONG_QUAD: E/W
    """
    try:
        d = _parse_float(deg) or 0.0
        m = _parse_float(minutes) or 0.0
        s = _parse_float(sec) or 0.0
        decimal = d + (m / 60.0) + (s / 3600.0)
        if quad and str(quad).strip().upper() in ("S", "W"):
            decimal = -decimal
        return decimal
    except Exception:
        return None


def _classify_medium(val) -> str:
    if not val:
        return Medium.UNKNOWN.value
    v = str(val).lower()
    if any(w in v for w in ["water", "river", "lake", "stream", "ocean", "bay", "sea", "creek"]):
        return Medium.WATER.value
    if any(w in v for w in ["air", "atmosphere"]):
        return Medium.AIR.value
    if any(w in v for w in ["land", "ground", "soil"]):
        return Medium.LAND.value
    return Medium.UNKNOWN.value


def _read_sheet(wb, sheet_name: str) -> dict[str, dict]:
    """
    Read a sheet into a dict keyed by SEQNOS.
    Returns {seqnos_str: {col: value, ...}}
    """
    if sheet_name not in wb.sheetnames:
        logger.warning("Sheet '%s' not found in workbook", sheet_name)
        return {}

    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)

    try:
        headers = [str(h).strip() if h is not None else f"col_{i}"
                   for i, h in enumerate(next(rows))]
    except StopIteration:
        return {}

    result = {}
    for row in rows:
        record = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
        seqnos = record.get("SEQNOS")
        if seqnos is not None:
            result[str(seqnos)] = record

    return result


class NRCIngestor(BaseIngestor):
    source = IncidentSource.NRC

    def __init__(self, db, data_dir: str, batch_size: int = 500):
        super().__init__(db, batch_size)
        self.data_dir = Path(data_dir)

    def fetch_records(self) -> Iterator[dict]:
        """
        For each xlsx file, join CALLS + INCIDENT_COMMONS + MATERIAL_INVOLVED
        on SEQNOS and yield one merged dict per incident.
        """
        if not HAS_OPENPYXL:
            logger.error("openpyxl not installed. Run: pip install openpyxl")
            return

        xlsx_files = sorted(self.data_dir.glob("nrc_*.xlsx"))

        if not xlsx_files:
            logger.warning(
                "No NRC xlsx files found in %s. "
                "Download from https://nrc.uscg.mil/FOIAdata.aspx and name them nrc_YYYY.xlsx",
                self.data_dir,
            )
            return

        for path in xlsx_files:
            logger.info("Reading %s", path.name)
            try:
                yield from self._process_file(path)
            except Exception as exc:
                logger.error("Failed to read %s: %s", path, exc)

    def _process_file(self, path: Path) -> Iterator[dict]:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        logger.info("  Sheets: %s", wb.sheetnames)

        # Load the three sheets we need
        calls = _read_sheet(wb, "CALLS")
        commons = _read_sheet(wb, "INCIDENT_COMMONS")
        materials = _read_sheet(wb, "MATERIAL_INVOLVED")

        wb.close()

        logger.info(
            "  Loaded: CALLS=%d INCIDENT_COMMONS=%d MATERIAL_INVOLVED=%d",
            len(calls), len(commons), len(materials),
        )

        # CALLS is the base — iterate over every incident
        for seqnos, call in calls.items():
            common = commons.get(seqnos, {})
            material = materials.get(seqnos, {})

            # Merge all three into one flat record, prefixed to avoid collisions
            merged = {"SEQNOS": seqnos}
            for k, v in call.items():
                merged[f"CALL_{k}"] = v
            for k, v in common.items():
                merged[f"COMMON_{k}"] = v
            for k, v in material.items():
                merged[f"MAT_{k}"] = v

            yield merged

    def normalize(self, raw: dict) -> dict | None:
        """Map a merged NRC record to the unified incident schema."""

        # --- Coordinates ---
        # INCIDENT_COMMONS stores coordinates in DMS format
        lat = _dms_to_decimal(
            raw.get("COMMON_LAT_DEG"),
            raw.get("COMMON_LAT_MIN"),
            raw.get("COMMON_LAT_SEC"),
            raw.get("COMMON_LAT_QUAD"),
        )
        lng = _dms_to_decimal(
            raw.get("COMMON_LONG_DEG"),
            raw.get("COMMON_LONG_MIN"),
            raw.get("COMMON_LONG_SEC"),
            raw.get("COMMON_LONG_QUAD"),
        )

        if lat is None or lng is None:
            return None
        if lat == 0.0 and lng == 0.0:
            return None
        if not (-90 <= lat <= 90) or not (-180 <= lng <= 180):
            return None

        # --- Source ID ---
        source_id = str(raw.get("SEQNOS", ""))
        if not source_id:
            return None

        # --- Material and medium ---
        material = raw.get("MAT_NAME_OF_MATERIAL")
        quantity = _parse_float(raw.get("MAT_AMOUNT_OF_MATERIAL"))
        quantity_unit = raw.get("MAT_UNIT_OF_MEASURE")

        # Medium: check if material reached water first, then fall back to description
        reached_water = str(raw.get("MAT_IF_REACHED_WATER", "")).upper()
        if reached_water == "YES":
            medium = Medium.WATER.value
        else:
            medium = _classify_medium(raw.get("COMMON_INCIDENT_LOCATION", ""))

        # --- Date ---
        incident_date = _parse_date(raw.get("COMMON_INCIDENT_DATE_TIME")) or \
                        _parse_date(raw.get("CALL_DATE_TIME_RECEIVED"))

        # --- Location ---
        city = raw.get("COMMON_LOCATION_NEAREST_CITY") or raw.get("CALL_RESPONSIBLE_CITY")
        state = raw.get("COMMON_LOCATION_STATE") or raw.get("CALL_RESPONSIBLE_STATE")
        address = raw.get("COMMON_LOCATION_STREET1")

        return {
            "id": uuid.uuid4(),
            "source": IncidentSource.NRC.value,
            "source_id": source_id,
            "incident_type": IncidentType.SPILL.value,
            "severity": None,
            "material": str(material) if material else None,
            "quantity": quantity,
            "quantity_unit": str(quantity_unit) if quantity_unit else None,
            "medium": medium,
            "facility_name": None,
            "responsible_party": raw.get("CALL_RESPONSIBLE_COMPANY"),
            "address": str(address) if address else None,
            "city": str(city) if city else None,
            "state": str(state)[:2] if state else None,
            "zip_code": raw.get("COMMON_LOCATION_ZIP"),
            "lat": lat,
            "lng": lng,
            "location": f"SRID=4326;POINT({lng} {lat})",
            "incident_date": incident_date,
            "raw": {k: str(v) if v is not None else None for k, v in raw.items()}
        }
